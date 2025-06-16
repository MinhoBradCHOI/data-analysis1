def create_achv_rate_sheet(ws4, indexed_table_df, target_wafer_ids):
    from openpyxl.styles import Font, Alignment
    import pandas as pd

    rate_cols = [f"Rate_{w}" for w in target_wafer_ids]
    col_labels = ['모듈', 'ITEM 개수'] + [f"Achv. rate{w}" for w in target_wafer_ids]

    for col_index, col_name in enumerate(col_labels, start=1):
        cell = ws4.cell(row=1, column=col_index, value=col_name)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True, size=9)

    row_labels = ['Total', 'ACTIVE', 'BCAT', 'Cell_TR', 'DCC', 'GBC', 'GBL', 'GP', 'BP', 'DCCP', 'CAP', 'BEOL', 'NMOS', 'PMOS']
    layer_mapping = {'BP': ['BP', 'BP_2F'], 'DCCP': ['DCCP', 'DCCP_2F']}

    for row_idx, layer in enumerate(row_labels, start=2):
        if layer == 'Total':
            subset = indexed_table_df[rate_cols]
        elif layer in layer_mapping:
            subset = indexed_table_df[indexed_table_df['Layer'].isin(layer_mapping[layer])][rate_cols]
        else:
            subset = indexed_table_df[indexed_table_df['Layer'] == layer][rate_cols]

        total_count = len(subset)
        ws4.cell(row=row_idx, column=1, value=layer)
        ws4.cell(row=row_idx, column=2, value=total_count)

        for col_idx, rate_col in enumerate(rate_cols, start=3):
            if total_count > 0:
                achieved = subset[rate_col].apply(pd.to_numeric, errors='coerce') > 0.8
                rate_percent = (achieved.sum() / total_count) * 100
            else:
                rate_percent = 0

            cell = ws4.cell(row=row_idx, column=col_idx, value=round(rate_percent))
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)

    for col in range(1, len(col_labels) + 1):
        ws4.column_dimensions[chr(64 + col)].width = 12

    for row in ws4.iter_rows(min_row=1, max_row=ws4.max_row, min_col=1, max_col=len(col_labels)):
        for cell in row:
            if cell.row in [1, 2] or cell.column == 1:
                cell.font = Font(bold=True, size=9)
            else:
                cell.font = Font(size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
