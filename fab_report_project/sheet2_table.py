def create_sheet2(ws2, indexed_table_df):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font, Alignment

    # ▶ 데이터프레임 삽입
    for row in dataframe_to_rows(indexed_table_df, index=False, header=True):
        if any(cell != "" and cell is not None for cell in row):
            ws2.append(row)

    # ▶ 스타일 지정
    for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row):
        for cell in row:
            col_name = ws2.cell(row=1, column=cell.column).value
            cell.font = Font(size=9, bold=(cell.row == 1))
            if cell.row == 1:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                continue

            if isinstance(col_name, str) and col_name.startswith("Rate_"):
                try:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.0%'
                        cell.value = round(cell.value, 3)
                except:
                    pass
            elif isinstance(col_name, str) and col_name.startswith("Tol_"):
                try:
                    if isinstance(cell.value, float):
                        if abs(cell.value) >= 1e4 or abs(cell.value) < 1e-4:
                            cell.number_format = '0.00E+00'
                        else:
                            cell.number_format = '0.00'
                except:
                    pass

            if col_name not in ['ITEM', 'ITEM_ID']:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # ▶ 열 너비 모두 10으로 고정
    for col in ws2.iter_cols(min_row=1, max_row=1):
        for cell in col:
            ws2.column_dimensions[cell.column_letter].width = 10

    # ▶ 빈 두 번째 행 제거
    ws2.delete_rows(2, 1)
