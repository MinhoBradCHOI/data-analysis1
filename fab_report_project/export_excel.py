import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import numpy as np

from config import target_root_lot_id, target_wafer_ids, save_dir
from preprocess import preprocess_data
from sheet1_merged import create_merged_sheet
from sheet2_table import create_sheet2
from sheet3_summary import create_summary_sheet
from sheet4_achv_rate import create_achv_rate_sheet

def export_excel():
    merged_df = preprocess_data()
    wb = Workbook()

    # Sheet1
    ws1 = wb.active
    ws1.title = "MergedSheet"
    create_merged_sheet(ws1, merged_df)

    # Sheet2 ~ Sheet4 전처리
    from copy import deepcopy
    empty_table_df = pd.DataFrame('', index=range(98), columns=['Layer', 'ITEM', 'ITEM_ID', 'MV T/G MED', 'MV T/G Tol.'])

    layer_values = ['ACTIVE'] * 15 + ['BCAT'] * 4 + ['Cell_TR'] * 8 + ['DCC'] * 2 + ['GBC'] * 9 + ['GBL'] * 3 + ['GP'] * 7 + ['BP'] * 3 + ['BP_2F'] * 5 + ['DCCP'] * 4 + ['DCCP_2F'] * 2 + ['CAP'] * 4 + ['BEOL'] * 14 + ['NMOS'] * 7 + ['PMOS'] * 10
    item_values = ['RWL', 'CWL', 'Cs', 'RBL', 'CBL','VTS', 'IDR', 'SWS', 'Cov', 'VFB'] + [f'ITEM{i}' for i in range(1, 88)]
    mv_tg_med_values = ['176000', '27.9', '2.7', '154200','28.4', '0.65', '2', '122', '-0.43', '1.9'] + [str(i) for i in range(1, 88)]
    mv_tg_tol_values = ['17600.0', '2.79', '0.27', '15420.0', '2.84', '0.065', '0.2', '12.2', '-0.043', '0.19'] + [str(round(i * 0.1 + 0.1, 2)) for i in range(87)]

    empty_table_df.loc[1:, 'Layer'] = layer_values
    empty_table_df.loc[1:, 'ITEM'] = item_values
    empty_table_df.loc[1:, 'ITEM_ID'] = item_values
    empty_table_df.loc[1:, 'MV T/G MED'] = mv_tg_med_values
    empty_table_df.loc[1:, 'MV T/G Tol.'] = mv_tg_tol_values

    med_cols, tol_cols, rate_cols = [], [], []

    for wafer_id in target_wafer_ids:
        filtered = merged_df[(merged_df['root_lot_id'] == target_root_lot_id) & (merged_df['wafer_id'] == wafer_id)]
        med_list, tol_list, rate_list = [], [], []

        for idx, row in empty_table_df.loc[1:, :].iterrows():
            item_id = row['ITEM_ID']
            mv_val = pd.to_numeric(row['MV T/G MED'], errors='coerce')
            raw_value = filtered.get(item_id)
            if isinstance(raw_value, pd.Series):
                values = pd.to_numeric(raw_value, errors='coerce').dropna()
            else:
                values = pd.Series([raw_value])
                values = pd.to_numeric(values, errors='coerce').dropna()

            if not values.empty and pd.notnull(mv_val):
                med_val = float(np.median(values))
                tol_val = float(np.std(values))
                if med_val < mv_val and mv_val != 0:
                    rate = med_val / mv_val
                elif med_val != 0:
                    rate = mv_val / med_val
                else:
                    rate = '-'
            else:
                med_val = tol_val = rate = '-'

            med_list.append(med_val)
            tol_list.append(tol_val)
            rate_list.append(round(rate, 4) if isinstance(rate, float) else rate)

        empty_table_df.loc[1:, f'MED_{wafer_id}'] = med_list
        empty_table_df.loc[1:, f'Tol_{wafer_id}'] = tol_list
        empty_table_df.loc[1:, f'Rate_{wafer_id}'] = rate_list
        med_cols.append((f'MED_{wafer_id}', f'MED_{wafer_id}'))
        tol_cols.append((f'Tol_{wafer_id}', f'Tol_{wafer_id}'))
        rate_cols.append((f'Rate_{wafer_id}', f'Rate_{wafer_id}'))

    empty_table_df.rename(columns=dict(med_cols + tol_cols + rate_cols), inplace=True)
    indexed_table_df = empty_table_df.copy()
    indexed_table_df.insert(0, 'Index', [''] + list(range(1, 98)))

    # Sheet2
    ws2 = wb.create_sheet(title=target_root_lot_id)
    create_sheet2(ws2, indexed_table_df)

    # Sheet3
    ws3 = wb.create_sheet(title="Summary")
    create_summary_sheet(ws3, indexed_table_df, target_wafer_ids)

    # Sheet4
    ws4 = wb.create_sheet(title="Achv_rate")
    create_achv_rate_sheet(ws4, indexed_table_df, target_wafer_ids)

    # 파일 저장
    today_str = datetime.today().strftime("%Y%m%d")
    base_name = f"{target_root_lot_id}_{'_'.join(['W' + w for w in target_wafer_ids])}_{today_str}"
    file_name = base_name + ".xlsx"
    full_path = os.path.join(save_dir, file_name)
    counter = 1
    while os.path.exists(full_path):
        file_name = f"{base_name}({counter}).xlsx"
        full_path = os.path.join(save_dir, file_name)
        counter += 1

    wb.save(full_path)
    os.startfile(save_dir)
    return full_path
