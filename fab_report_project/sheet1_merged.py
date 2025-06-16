def create_merged_sheet(ws1, merged_df):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    merged_df_sorted = merged_df.sort_values(by='root_lot_id')
    for row in dataframe_to_rows(merged_df_sorted, index=False, header=True):
        ws1.append(row)

    # ✅ 안전한 테이블 범위 생성
    last_col_letter = get_column_letter(ws1.max_column)
    table_range = f"A1:{last_col_letter}{ws1.max_row}"

    table = Table(displayName="MergedData", ref=table_range)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False,
                                          showRowStripes=False, showColumnStripes=False)
    ws1.add_table(table)

    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
            if cell.row == 1:
                cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
